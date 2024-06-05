from openpyxl import load_workbook
import os
from datetime import datetime
import pandas as pd

def update_excel(client_file,match_sheet, document_number, doctype, date_time):
    # client_file = "SubBookUpload\Data\Sub book tracker_IIMI_Southern Region.xlsx"
    # client_file = input("Please enter the path to the client excel file: ")

    wb = load_workbook(client_file)
    file_name = os.path.basename(client_file)
    sheetnames = wb.sheetnames

    for sheetname in sheetnames:
        if match_sheet in sheetname:

            data_dict = {
                'NUMBER': [document_number],
                'Input ': [doctype],
                'Transmission  Date': [date_time]
            }
            print(f"Match found: {match_sheet} matches with {sheetname} Number : {data_dict['NUMBER']}, Docktype or input : {data_dict['Input ']}")


            sheet = wb[sheetname]


            start_col_dlt = 24 
            end_col_dlt = 37    

            for col in range(end_col_dlt, start_col_dlt - 1, -1):
                sheet.delete_cols(col)

            wb.save(client_file)
            break
    else:
        print(f"No match found for {match_sheet}")




def extract_data(opt_file):
    df = pd.read_csv(opt_file)
    extracted_data = df[["FileName", "DOCUMENT NUM", "DOCTYPE", "DATE-TIME"]]

    data_list = extracted_data.to_dict(orient="records")
    unique_filenames = set(record['FileName'] for record in data_list)
    
    filtered_records_by_filename = {}
    for filename in unique_filenames:
        filtered_records = [record for record in data_list if record['FileName'] == filename]
        filtered_records_by_filename[filename] = filtered_records
    

    return filtered_records_by_filename 


def main():
    opt_file = input("Please enter the path to the Output CSV file: ")
    client_file = input("Please enter the path to the client excel file: ")

    # opt_file = "SubBookUpload\Data\output.csv"
    # client_file = "SubBookUpload\Data\Sub book tracker_IIMI_Southern Region.xlsx"

    filtered_data = extract_data(opt_file)
    for filename, records in filtered_data.items():
        match_sheet = filename[:4]
        for record in records:
            document_number =  record['DOCUMENT NUM']
            doctype =  record['DOCTYPE']
            date_time =  record['DATE-TIME'][4:14]

            update_excel(client_file,match_sheet,document_number, doctype, date_time)


if __name__ == "__main__":
    today_date = datetime.now().strftime('%Y-%m-%d')
    target_date = datetime(year=2024, month=12, day=28)
    current_date = datetime.now()
    print(f'''
                                                █████ █████ ██████   ██████ █████      
                                                 ███   ███   ██████ ██████   ███       
                                                 ███   ███   ███  ███  ███   ███                   
                                                 ███   ███   ███       ███   ███    
                                                █████ █████ █████     █████ █████
                                                            V-0.0.1 (System Team)  
          
                                                   🎉🎉    Sub Book Delete  Bot   🎉🎉 
                                                              {today_date}                                                   
    ''')

    if current_date < target_date: 
        try:
            username = input('Enter username : ').lower()
            password = input('Enter password : ').lower()
            if username == 'iimi' and password == 'iimi':
                main()
                input("Press any key to close.")
            else:
                print("Username and password  incorrect 😢😢😢")

        except Exception as e:
            print(f"An error occurred: {e}")
            
    else:
        print("Update your software!")


# pyinstaller --onefile --icon=logo.ico BotMap.py